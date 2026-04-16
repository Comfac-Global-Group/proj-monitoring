#!/usr/bin/env python3
"""
XLSX Migration Script for Project Monitoring Log (PML)
Converts the 'OTHER MATTERS' sheet from the Plant Operations Meeting spreadsheet
into the app's native JSON import format.

Usage:
    python3 scripts/xlsx-to-json.py \
        "/path/to/PLANT OPERATIONS MEETING NEW_ 2024 - 2026.xlsx" \
        > other-matters-import.json

Mapping:
    DONE                -> status (closed if True, else open)
    ITEMS / PROJ        -> title
    DETAILS             -> details
    OWNER               -> actions[].owner
    ACTION TO BE TAKEN  -> actions[].text  (with embedded yymmdd log entries)
    ACTION DUE DATE     -> actions[].due_date
    PROJ DUE DATE       -> project_due_date
    ISSUE / CAUSE OF DELAY -> actions[].issue
    REMARKS             -> notes
    LINK                -> appended to notes
"""

import sys
import json
import re
import uuid
from datetime import datetime
from openpyxl import load_workbook


def generate_id():
    return f"id-{uuid.uuid4().hex[:9]}"


def get_timestamp():
    now = datetime.now()
    return f"{now.year % 100:02d}{now.month:02d}{now.day:02d}-{now.hour:02d}{now.minute:02d}{now.second:02d}"


def parse_date(value):
    """Return YYYY-MM-DD string from datetime or string, or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Try common formats
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def parse_log_entries(text):
    """Extract yymmdd-prefixed lines from action text."""
    entries = []
    if not text:
        return entries
    # Pattern: 6 digits, optional whitespace, optional dash/em-dash, optional whitespace, then text
    log_pattern = re.compile(r"^(\d{6})\s*[-–]?\s*(.*)$")
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        m = log_pattern.match(line)
        if m:
            entries.append({
                "date": m.group(1),
                "text": m.group(2).strip()
            })
    return entries


def convert_sheet(wb_path, sheet_name="OTHER MATTERS"):
    wb = load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=5, values_only=True))

    projects = []
    timestamp = get_timestamp()

    for row in rows:
        if not row or len(row) < 4:
            continue

        done_flag = row[0]
        title = row[3]
        details = row[4] or ""
        owner = row[5] or ""
        action_text = row[6] or ""
        action_due_date = parse_date(row[7])
        proj_due_date = parse_date(row[8])
        issue = row[9] or ""
        remarks = row[10] or ""
        link = row[11] or ""

        if not title:
            continue

        notes = str(remarks) if remarks else ""
        if link:
            if notes:
                notes += "\n"
            notes += str(link)

        status = "closed" if done_flag is True else "open"

        actions = []
        if action_text:
            actions.append({
                "id": generate_id(),
                "text": str(action_text).strip(),
                "due_date": action_due_date or "",
                "owner": str(owner).strip() if owner else "",
                "issue": str(issue).strip() if issue else "",
                "comments": [],
                "log_entries": parse_log_entries(str(action_text)),
                "created_at": timestamp,
                "updated_at": timestamp
            })

        project = {
            "id": generate_id(),
            "title": str(title).strip(),
            "status": status,
            "details": str(details).strip() if details else "",
            "notes": notes,
            "project_due_date": proj_due_date or "",
            "actions": actions,
            "notesComments": [],
            "created_at": timestamp,
            "updated_at": timestamp
        }
        projects.append(project)

    return {"projects": projects}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: xlsx-to-json.py <path-to-xlsx> [sheet-name]", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else "OTHER MATTERS"
    data = convert_sheet(xlsx_path, sheet)
    print(json.dumps(data, indent=2, ensure_ascii=False))
